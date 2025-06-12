import pandas as pd
import os
import sys
import numpy as np
from openpyxl import Workbook, load_workbook

def is_numeric(value):
    """Check if a value is numeric (either integer or float) and not NaN"""
    if pd.isna(value):  # Check for NaN values
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        # Remove any whitespace
        value = value.strip()
        # Check if string can be converted to a number
        try:
            float(value)
            return True
        except ValueError:
            return False
    return False

def safe_convert_to_int(value):
    """Safely convert a value to an integer string representation"""
    try:
        # First check if it's NaN
        if pd.isna(value):
            return str(value)
        
        # Then try to convert to float first, then to int
        float_val = float(value)
        # Only convert to int if it's a whole number
        if float_val.is_integer():
            return str(int(float_val))
        return str(float_val)
    except:
        # If any conversion fails, just return string representation
        return str(value)

def extract_io_list(input_file):
    """
    Extracts data from the 'IO LIST' worksheet and creates a new Excel file
    with separate worksheets for each rack/slot combination.
    """
    # Load the Excel file
    print(f"Opening Excel file for IO extraction: {input_file}")
    try:
        # Read the IO LIST worksheet
        df = pd.read_excel(input_file, sheet_name="IO LIST")
        print("Successfully loaded 'IO LIST' worksheet")
        print(f"Loaded {len(df)} rows from worksheet")
        
        # Column indices (0-indexed)
        column_indices = {
            'rack': 2,     # Column C
            'slot': 3,     # Column D
            'io_point': 4, # Column E
            'col_h': 7,    # Column H
            'col_k': 10,   # Column K
            'col_q': 16    # Column Q
        }
        
        # Create a new workbook
        output_workbook = Workbook()
        # Remove default sheet
        output_workbook.remove(output_workbook.active)
        
        # Count of rows and worksheets processed
        total_rows = 0
        worksheets_created = 0
        skipped_rows = 0
        skipped_rack_slot_combinations = 0
        
        # Extract data from each row and organize by rack and slot
        rack_slot_data = {}
        
        # Process each row
        for index, row in df.iterrows():
            # Get values for the current row
            rack = row.iloc[column_indices['rack']]
            slot = row.iloc[column_indices['slot']]
            io_point = row.iloc[column_indices['io_point']]
            
            # Debug output for problematic rows
            if index < 5 or pd.isna(rack) or pd.isna(slot) or pd.isna(io_point):
                print(f"Debug Row {index}: Rack={rack}, Slot={slot}, IO Point={io_point}")
            
            # Check if rack and slot are numeric and not NaN
            if not (is_numeric(rack) and is_numeric(slot)):
                skipped_rack_slot_combinations += 1
                continue
                
            # Check if io_point is numeric and not NaN
            if not is_numeric(io_point):
                skipped_rows += 1
                continue
                
            # Convert to strings for consistency in dictionary keys
            rack_str = safe_convert_to_int(rack)
            slot_str = safe_convert_to_int(slot)
            io_point_str = safe_convert_to_int(io_point)
            
            # Get other column values
            col_h_val = row.iloc[column_indices['col_h']]
            col_k_val = row.iloc[column_indices['col_k']]
            col_q_val = row.iloc[column_indices['col_q']]
            
            # Create key for this rack and slot combination
            key = f"{rack_str}_{slot_str}"
            
            # Initialize list for this rack/slot if it doesn't exist
            if key not in rack_slot_data:
                rack_slot_data[key] = []
                
            # Add row data to the appropriate rack/slot list
            rack_slot_data[key].append([
                rack_str, 
                slot_str, 
                io_point_str,
                col_h_val,
                col_k_val,
                col_q_val
            ])
            
            total_rows += 1
        
        # Create worksheets for each valid rack/slot combination
        for key, data in rack_slot_data.items():
            rack_str, slot_str = key.split('_')
            
            # Create worksheet name
            worksheet_name = f"Rack {rack_str} Slot {slot_str}"
            # Make sure worksheet name is valid (max 31 chars, no special chars)
            worksheet_name = worksheet_name[:31].replace(':', '_')
            print(f"Creating worksheet: {worksheet_name}")
            
            # Create a new worksheet
            ws = output_workbook.create_sheet(title=worksheet_name)
            
            # Add headers
            headers = ['Rack', 'Slot', 'IO Point', 'Column H', 'Column K', 'Column Q']
            ws.append(headers)
            
            # Add data rows
            for row_data in data:
                ws.append(row_data)
                
            worksheets_created += 1
        
        # Save the output workbook
        output_file = os.path.splitext(input_file)[0] + "_processed.xlsx"
        output_workbook.save(output_file)
        print(f"Successfully created output file: {output_file}")
        print(f"Summary: {total_rows} rows processed, {worksheets_created} worksheets created")
        print(f"Skipped {skipped_rack_slot_combinations} rack/slot combinations (non-numeric values)")
        print(f"Skipped {skipped_rows} rows (non-numeric IO points)")
        
        return output_file
    
    except Exception as e:
        print(f"Error processing Excel file: {e}")
        import traceback
        traceback.print_exc()
        return None

def convert_excel_to_csv(excel_file, output_dir=None):
    """
    Converts each worksheet in an Excel file to a separate CSV file.
    Returns the directory containing the CSV files and a list of generated filenames.
    """
    print(f"Opening Excel file for CSV conversion: {excel_file}")
    
    try:
        # Get all sheet names from the Excel file
        wb = load_workbook(filename=excel_file, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        
        if not sheet_names:
            print("No worksheets found in the Excel file.")
            return None, []
            
        print(f"Found {len(sheet_names)} worksheets")
        
        # Create output directory
        if output_dir is None:
            base_name = os.path.splitext(os.path.basename(excel_file))[0]
            output_dir = os.path.join(os.path.dirname(excel_file), f"{base_name}_csv")
        
        # Create directory if it doesn't exist
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # List to store generated CSV filenames
        generated_files = []
            
        # Process each worksheet
        for sheet_name in sheet_names:
            print(f"Processing worksheet: {sheet_name}")
            
            # Read the worksheet
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            if df.empty:
                print(f"  Sheet '{sheet_name}' is empty. Skipping.")
                continue
                
            # Create safe filename
            safe_sheet_name = sheet_name.replace(" ", "_").replace(":", "_")
            csv_file = os.path.join(output_dir, f"{safe_sheet_name}.csv")
            
            # Save as CSV
            df.to_csv(csv_file, index=False)
            print(f"  Saved as: {csv_file}")
            
            # Add to list of generated files
            generated_files.append(f"{safe_sheet_name}.csv")
        
        print(f"\nCSV files saved to directory: {output_dir}")
        
        # Create RackList.csv with all filenames
        create_rack_list_csv(output_dir, generated_files)
        
        return output_dir, generated_files
        
    except Exception as e:
        print(f"Error processing Excel file: {e}")
        import traceback
        traceback.print_exc()
        return None, []

def create_rack_list_csv(output_dir, filenames):
    """
    Creates a CSV file containing a list of all generated CSV filenames.
    """
    try:
        rack_list_path = os.path.join(output_dir, "RackList.csv")
        print(f"\nCreating RackList.csv with {len(filenames)} filenames...")
        
        # Create DataFrame with filenames
        df = pd.DataFrame(filenames, columns=["Filename"])
        
        # Save to CSV
        df.to_csv(rack_list_path, index=False)
        print(f"RackList.csv created successfully at: {rack_list_path}")
        return rack_list_path
    except Exception as e:
        print(f"Error creating RackList.csv: {e}")
        import traceback
        traceback.print_exc()
        return None

# def main():
#     """Main function that handles command line arguments and user interaction"""
#     # Display welcome message
#     print("Excel File Processor")
#     print("-------------------")
#     print("1. Extract IO List from Excel file")
#     print("2. Convert Excel worksheets to CSV files")
#     print("3. Extract IO List and then convert to CSV")
#     print("4. Exit")
    
#     choice = input("\nEnter your choice (1-4): ")
    
#     if choice == "4":
#         print("Exiting program.")
#         sys.exit(0)
    
#     # Get input file
#     if len(sys.argv) > 1:
#         input_file = sys.argv[1]
#     else:
#         input_file = input("\nEnter the path to the Excel file: ")
    
#     # Check if file exists
#     if not os.path.exists(input_file):
#         print("File not found. Please check the file path.")
#         return
    
#     # Process based on user choice
#     if choice == "1":
#         # Extract IO List only
#         output_file = extract_io_list(input_file)
#         if output_file:
#             print(f"\nProcess completed. Output saved to {output_file}")
#         else:
#             print("\nProcess failed. Please check the errors above.")
            
#     elif choice == "2":
#         # Convert to CSV only
#         output_dir = None
#         if len(sys.argv) > 2:
#             output_dir = sys.argv[2]
#         else:
#             output_dir_input = input("Enter the output directory (leave blank for default): ")
#             if output_dir_input.strip() != "":
#                 output_dir = output_dir_input
        
#         result_dir, generated_files = convert_excel_to_csv(input_file, output_dir)
        
#         if result_dir:
#             print("\nProcess completed successfully.")
#             print(f"\nGenerated {len(generated_files)} CSV files plus RackList.csv")
#             print("\nTo use these CSV files in your Typst document:")
#             print(f"1. Save the io_tables.typ library to your Typst project")
#             print(f"2. Add this to your main Typst document:")
#             print(f"   #import \"io_tables.typ\": io-tables")
#             print(f"   #io-tables.generate-io-tables(\"{result_dir}\")")
#         else:
#             print("\nProcess failed.")
            
#     elif choice == "3":
#         # Extract IO List and then convert to CSV
#         processed_file = extract_io_list(input_file)
#         if not processed_file:
#             print("\nIO List extraction failed. Stopping process.")
#             return
            
#         print("\nNow converting the processed file to CSV...")
        
#         output_dir = input("Enter the output directory for CSV files (leave blank for default): ")
#         if output_dir.strip() == "":
#             output_dir = None
            
#         result_dir, generated_files = convert_excel_to_csv(processed_file, output_dir)
        
#         if result_dir:
#             print("\nProcess completed successfully.")
#             print(f"\nGenerated {len(generated_files)} CSV files plus RackList.csv")
#             print("\nTo use these CSV files in your Typst document:")
#             print(f"1. Save the io_tables.typ library to your Typst project")
#             print(f"2. Add this to your main Typst document:")
#             print(f"   #import \"io_tables.typ\": io-tables")
#             print(f"   #io-tables.generate-io-tables(\"{result_dir}\")")
#         else:
#             print("\nCSV conversion failed.")
#     else:
#         print("Invalid choice. Please run the program again and select a valid option.")

# if __name__ == "__main__":
#     main()